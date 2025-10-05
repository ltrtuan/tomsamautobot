# actions/write_csv_action.py

from controllers.actions.base_action import BaseAction
from models.global_variables import GlobalVariables
import os
import re
import random
import string
import csv

try:
    from openpyxl import load_workbook, Workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("[WRITE_CSV] Warning: openpyxl not installed. Excel files (.xlsx, .xls) are not supported.")

class WriteCsvAction(BaseAction):
    """Action to write content to next empty row in CSV/Excel file"""
    
    def prepare_play(self):
        """Execute write csv file action after delay"""
        if self.should_stop():
            return
        
        # Get file path (priority: variable > browse)
        file_path_var = self.params.get("file_path_variable", "").strip()
        file_path = self.params.get("file_path", "").strip()
        
        # Determine final file path
        if file_path_var:
            globals_var = GlobalVariables()
            final_path = globals_var.get(file_path_var, "")
            print(f"[WRITE_CSV] Using variable '{file_path_var}': {final_path}")
        else:
            final_path = file_path
            print(f"[WRITE_CSV] Using browse path: {final_path}")
        
        # Validate file path
        if not final_path:
            print("[WRITE_CSV] No file path provided. Skipping.")
            return
        
        # ✅ CHECK: File có tồn tại không - SKIP nếu không tồn tại
        if not os.path.exists(final_path):
            print(f"[WRITE_CSV] File does not exist: {final_path}. Skipping write operation.")
            return
        
        # Get content list
        content_list_str = self.params.get("content", "").strip()
        
        if not content_list_str:
            print("[WRITE_CSV] No content to write. Skipping.")
            return
        
        # ✅ PARSE: Split by semicolon - mỗi item là 1 column
        content_items = [item.strip() for item in content_list_str.split(";") if item.strip()]
        
        if not content_items:
            print("[WRITE_CSV] No valid content items. Skipping.")
            return
        
        # ✅ PROCESS ALL: Xử lý pattern cho TẤT CẢ items
        processed_items = []
        for item in content_items:
            processed_text = self._process_text(item)
            processed_items.append(processed_text)
        
        # ✅ WRITE TO FILE: Phân biệt CSV và Excel
        file_extension = os.path.splitext(final_path)[1].lower()
        
        if file_extension in ['.csv', '.txt']:
            self._write_to_csv(final_path, processed_items)
        elif file_extension in ['.xlsx', '.xls']:
            if EXCEL_SUPPORT:
                self._write_to_excel(final_path, processed_items)
            else:
                print(f"[WRITE_CSV] Excel support not available. Please install openpyxl: pip install openpyxl")
        else:
            print(f"[WRITE_CSV] Unsupported file format: {file_extension}")
    
    def _write_to_csv(self, file_path, processed_items):
        """Write to CSV file"""
        try:
            # Đọc tất cả rows hiện tại
            existing_rows = []
            
            try:
                with open(file_path, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.reader(f)
                    existing_rows = list(reader)
            except Exception as e:
                print(f"[WRITE_CSV] Could not read CSV file (may be empty): {e}")
                existing_rows = []
            
            # ✅ XÓA CÁC ROW HOÀN TOÀN RỖNG Ở CUỐI
            while existing_rows and all(cell.strip() == '' for cell in existing_rows[-1]):
                existing_rows.pop()
                print(f"[WRITE_CSV] Removed empty row at end")
            
            # ✅ ĐẢM BẢO SỐ COLUMNS KHỚP
            if existing_rows:
                max_cols = max(len(row) for row in existing_rows)
                
                if len(processed_items) < max_cols:
                    processed_items.extend([''] * (max_cols - len(processed_items)))
                
                for row in existing_rows:
                    if len(row) < max(max_cols, len(processed_items)):
                        row.extend([''] * (max(max_cols, len(processed_items)) - len(row)))
            
            # ✅ THÊM ROW MỚI
            existing_rows.append(processed_items)
            
            # ✅ GHI LẠI TOÀN BỘ FILE
            with open(file_path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(existing_rows)
            
            print(f"[WRITE_CSV] Successfully wrote {len(processed_items)} column(s) to row {len(existing_rows)} in: {file_path}")
            for idx, content in enumerate(processed_items, 1):
                print(f"[WRITE_CSV]   Column {idx}: {content}")
            
        except Exception as e:
            print(f"[WRITE_CSV] Error writing to CSV file: {e}")
            import traceback
            traceback.print_exc()
    
    def _write_to_excel(self, file_path, processed_items):
        """Write to Excel file"""
        try:
            # Load workbook
            try:
                wb = load_workbook(file_path)
                ws = wb.active
            except Exception as e:
                print(f"[WRITE_CSV] Could not load Excel file: {e}")
                return
            
            # ✅ TÌM ROW RỖNG ĐẦU TIÊN (xóa các row hoàn toàn rỗng ở cuối)
            max_row = ws.max_row
            
            # Đếm ngược từ cuối để tìm row không rỗng cuối cùng
            last_non_empty_row = 0
            for row_idx in range(max_row, 0, -1):
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
                # Kiểm tra xem row có ít nhất 1 cell không rỗng
                if any(val is not None and str(val).strip() != '' for val in row_values):
                    last_non_empty_row = row_idx
                    break
            
            # Row để ghi mới = row sau row không rỗng cuối cùng
            new_row_idx = last_non_empty_row + 1
            
            # ✅ GHI DỮ LIỆU VÀO ROW MỚI
            for col_idx, value in enumerate(processed_items, start=1):
                ws.cell(row=new_row_idx, column=col_idx, value=value)
            
            # ✅ XÓA CÁC ROW RỖNG (nếu có) giữa last_non_empty_row và row mới ghi
            # (Trong trường hợp này không cần vì ta đã ghi vào row ngay sau row không rỗng)
            
            # ✅ LƯU FILE
            wb.save(file_path)
            wb.close()
            
            print(f"[WRITE_CSV] Successfully wrote {len(processed_items)} column(s) to row {new_row_idx} in: {file_path}")
            for idx, content in enumerate(processed_items, 1):
                print(f"[WRITE_CSV]   Column {idx}: {content}")
            
        except Exception as e:
            print(f"[WRITE_CSV] Error writing to Excel file: {e}")
            import traceback
            traceback.print_exc()
    
    def _process_text(self, text):
        """Process special text formats (giống write_txt_action)"""
        # Replace with variable value
        globals_var = GlobalVariables()
        
        # Pattern for <VARIABLE>
        def replace_variable(match):
            var_name = match.group(1)
            value = globals_var.get(var_name, None)
            
            if value is not None:
                print(f"[WRITE_CSV] Replaced <{var_name}> with '{value}'")
                return str(value)
            else:
                print(f"[WRITE_CSV] WARNING: Variable '{var_name}' not found!")
                return f"<{var_name}>"
        
        text = re.sub(r'<([^>]+)>', replace_variable, text)
        
        # Pattern: [1-10:CN] - random 1-10 chars/numbers/both
        def replace_random_chars(match):
            min_val = int(match.group(1))
            max_val = int(match.group(2))
            char_type = match.group(3).upper()
            length = random.randint(min_val, max_val)
            
            if char_type == 'C':
                return ''.join(random.choices(string.ascii_letters, k=length))
            elif char_type == 'N':
                return ''.join(random.choices(string.digits, k=length))
            elif char_type == 'CN':
                return ''.join(random.choices(string.ascii_letters + string.digits, k=length))
            else:
                return match.group(0)
        
        text = re.sub(r'\[(\d+)-(\d+):([CNcn]+)\]', replace_random_chars, text)
        
        # Pattern: [1-10] - random number from 1 to 10
        def replace_random_number(match):
            min_val = int(match.group(1))
            max_val = int(match.group(2))
            return str(random.randint(min_val, max_val))
        
        text = re.sub(r'\[(\d+)-(\d+)\]', replace_random_number, text)
        
        return text
